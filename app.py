from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import pandas as pd

app = Flask(__name__)
CORS(app)

# File to store data
file_name = "VAT_Entries.xlsx"

# Create Excel file if it doesn't exist
if not os.path.exists(file_name):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "VAT Entries"
        ws.append(["Date", "Product Name", "Stock", "VAT Rate", "VAT Amount"])
        wb.save(file_name)
    except Exception as e:
        print(f"Error creating the Excel file: {e}")
        raise

# API endpoint to calculate VAT and save entry
@app.route('/calculate', methods=['POST'])
def calculate_vat():
    try:
        data = request.json
        date = data.get('date', datetime.now().strftime('%Y-%m-%d'))
        product_name = data['product_name']
        stock = float(data['stock'])
        vat_rate = float(data['vat_rate'])

        if stock < 0 or vat_rate < 0:
            return jsonify({"error": "Stock and VAT rate must be non-negative."}), 400

        vat_amount = stock * vat_rate

        # Save entry to Excel
        wb = load_workbook(file_name)
        ws = wb.active
        ws.append([date, product_name, stock, vat_rate, vat_amount])
        wb.save(file_name)

        # Calculate total VAT
        total_vat = sum(float(row[4]) for row in ws.iter_rows(min_row=2, values_only=True))

        return jsonify({"vat_amount": vat_amount, "total_vat": total_vat})

    except KeyError as e:
        return jsonify({"error": f"Missing data: {str(e)}"}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid data format: {str(e)}"}), 400
    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500

# API endpoint to fetch all entries
@app.route('/entries', methods=['GET'])
def get_entries():
    try:
        wb = load_workbook(file_name)
        ws = wb.active
        entries = [row for row in ws.iter_rows(min_row=2, values_only=True)]
        return jsonify(entries)
    
    except Exception as e:
        return jsonify({"error": f"Failed to fetch entries: {str(e)}"}), 500

# API endpoint to delete an entry
@app.route('/delete', methods=['POST'])
def delete_entry():
    try:
        row_index = int(request.json['row_index']) + 2  # Account for header row and 0-based index from JS

        wb = load_workbook(file_name)
        ws = wb.active

        # Delete the selected row
        ws.delete_rows(row_index)

        # Recalculate total VAT and update the last row
        total_vat = sum(float(row[4]) for row in ws.iter_rows(min_row=2, values_only=True))

        # Find the last row (which is the total VAT row) and delete it
        for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, values_only=True):
            if row[3] == "Total VAT:":
                ws.delete_rows(ws.max_row)

        # Add the new total VAT row
        ws.append(['', '', '', 'Total VAT:', total_vat])

        wb.save(file_name)
        return jsonify({"message": "Entry deleted successfully."})

    except IndexError as e:
        return jsonify({"error": "Invalid row index for deletion."}), 400
    except Exception as e:
        return jsonify({"error": f"An error occurred while deleting the entry: {str(e)}"}), 500

# API endpoint to edit an entry
@app.route('/edit', methods=['POST'])
def edit_entry():
    try:
        data = request.json
        row_index = int(data['row_index']) + 2

        wb = load_workbook(file_name)
        ws = wb.active
        ws[row_index][0].value = data['date']
        ws[row_index][1].value = data['product_name']
        ws[row_index][2].value = float(data['stock'])
        ws[row_index][3].value = float(data['vat_rate'])
        ws[row_index][4].value = float(data['stock']) * float(data['vat_rate'])

        wb.save(file_name)
        return jsonify({"message": "Entry updated successfully."})

    except IndexError as e:
        return jsonify({"error": "Invalid row index for update."}), 400
    except Exception as e:
        return jsonify({"error": f"An error occurred while editing the entry: {str(e)}"}), 500

# API endpoint to download Excel file
@app.route('/download', methods=['GET'])
def download_excel():
    try:
        return send_file(file_name, as_attachment=True)
    except Exception as e:
        return jsonify({"error": f"Failed to download file: {str(e)}"}), 500

# API endpoint to generate summary report
@app.route('/report', methods=['GET'])
def generate_report():
    try:
        df = pd.read_excel(file_name)
        summary = {
            "Total VAT Amount": df["VAT Amount"].sum(),
            "Total Entries": len(df),
        }
        return jsonify(summary)
    except Exception as e:
        return jsonify({"error": f"Failed to generate report: {str(e)}"}), 500

if __name__ == "__main__":
    app.run(debug=True)
